"""scripts/facebook-to-excel.py

Fetches Facebook Page posts via the Graph API and writes post URLs and labels
into the appropriate columns of data/ejc-data-tracker.xlsx.

Posts are matched to towns by extracting the visit number from the first line
of the post message and looking it up against the visitNumber column in the
Excel file.  Only posts that follow the "NJ Town #N:" or "New Jersey Town N:"
pattern are processed; all others are silently skipped.

Slot assignment:
    "Bonus Town Trivia" in the message first line → facebook2_label / facebook2_url
    All other matching posts                       → facebook1_label / facebook1_url

Labels are always normalized to:
    Main post:   "NJ Town #N: <Town Name>"        (town name from Excel, not caption)
    Bonus post:  "NJ Town #N: Bonus Town Trivia"

Cells that already contain data are never overwritten.
A timestamped backup of the Excel file is created before any changes are saved.

The Page Access Token can be supplied via the --token argument or the
FB_PAGE_TOKEN environment variable (env var preferred — keeps the token out of
shell history and process listings).

Usage:
    # Token via env var (recommended):
    FB_PAGE_TOKEN=your_token .venv/bin/python scripts/facebook-to-excel.py

    # Token via argument:
    .venv/bin/python scripts/facebook-to-excel.py --token YOUR_PAGE_TOKEN

    # Explicit page ID (skips auto-detection):
    .venv/bin/python scripts/facebook-to-excel.py --token YOUR_TOKEN --page-id PAGE_ID

    # Custom Excel path:
    .venv/bin/python scripts/facebook-to-excel.py --excel path/to/other.xlsx

How to get a Page Access Token:
    1. Go to https://developers.facebook.com/tools/explorer
    2. Create or select a Meta App
    3. Click "Get User Token" and grant pages_read_engagement + pages_show_list
    4. In the Explorer, call GET /me/accounts to find your page and copy its
       access_token — that is your Page Access Token.
    5. Optionally exchange for a long-lived (60-day) token via the Token Debugger.
"""

import argparse
import os
import re
import shutil
import sys
from datetime import date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit(
        "ERROR: openpyxl is not installed.\n"
        "       Activate the virtual environment and run: pip install openpyxl"
    )

try:
    import requests
except ImportError:
    sys.exit(
        "ERROR: requests is not installed.\n"
        "       Activate the virtual environment and run: pip install requests"
    )

# ── Repo paths ────────────────────────────────────────────────────────────────
_REPO_ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = _REPO_ROOT / "data" / "ejc-data-tracker.xlsx"

# ── Graph API ─────────────────────────────────────────────────────────────────
_GRAPH_VERSION = "v21.0"
_GRAPH_BASE    = f"https://graph.facebook.com/{_GRAPH_VERSION}"
_PAGE_SIZE     = 100   # maximum allowed by the API

# ── Caption matching ──────────────────────────────────────────────────────────
# Matches the start of a post first line like:
#   "NJ Town #93: Carteret"
#   "NJ Town 93: Carteret"
#   "New Jersey Town #93: Carteret"
#   "NJ Town #93: Carteret - Bonus Town Trivia"
# Captures the numeric visit number.
_TOWN_RE = re.compile(
    r"^(?:NJ|New\s+Jersey)\s+Town\s+#?(\d+)\s*[:\-,]",
    re.IGNORECASE,
)

# Detects "Bonus Town Trivia" anywhere in the first message line.
_BONUS_RE = re.compile(r"Bonus\s+Town\s+Trivia", re.IGNORECASE)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _first_line(text: str) -> str:
    """Return the first non-empty line of a multi-line string."""
    for line in text.splitlines():
        stripped = line.strip()
        if stripped:
            return stripped
    return ""


def _make_backup(excel_path: Path) -> Path:
    """
    Copy the Excel file to a date-stamped backup path.
    Does not overwrite a backup already created earlier the same day.
    Returns the backup path.
    """
    stamp  = date.today().strftime("%Y%m%d")
    backup = excel_path.with_name(f"ejc-data-tracker-backup-{stamp}.xlsx")
    if not backup.exists():
        shutil.copy2(excel_path, backup)
        print(f"Backup created : {backup}")
    else:
        print(f"Backup exists  : {backup}  (not overwritten)")
    return backup


def _graph_get(endpoint: str, params: dict) -> dict:
    """
    Perform a GET request against the Graph API and return the parsed JSON.
    Exits with a descriptive error on non-200 responses or API error objects.
    """
    url = f"{_GRAPH_BASE}/{endpoint.lstrip('/')}"
    resp = requests.get(url, params=params, timeout=30)
    try:
        data = resp.json()
    except ValueError:
        sys.exit(
            f"ERROR: Non-JSON response from Graph API (HTTP {resp.status_code}).\n"
            f"       URL: {url}"
        )

    if not resp.ok:
        err = data.get("error", {})
        sys.exit(
            f"ERROR: Graph API returned HTTP {resp.status_code}.\n"
            f"       Message : {err.get('message', 'unknown error')}\n"
            f"       Type    : {err.get('type', '')}\n"
            f"       Code    : {err.get('code', '')}\n"
            "\nCheck that your Page Access Token is valid and has the required permissions:\n"
            "  pages_read_engagement, pages_show_list"
        )

    if "error" in data:
        err = data["error"]
        sys.exit(
            f"ERROR: Graph API error in response body.\n"
            f"       Message : {err.get('message', 'unknown error')}\n"
            f"       Type    : {err.get('type', '')}\n"
            f"       Code    : {err.get('code', '')}"
        )

    return data


# ── Core logic ────────────────────────────────────────────────────────────────

def get_page_id(token: str) -> tuple[str, str]:
    """
    Auto-detect the Page ID and name from the token by calling /me/accounts.
    Returns (page_id, page_name).
    Exits if the token does not have access to exactly one page or if the API
    responds with an error.
    """
    data  = _graph_get("me/accounts", {"access_token": token, "fields": "id,name"})
    pages = data.get("data", [])

    if not pages:
        sys.exit(
            "ERROR: No Facebook Pages found for this token.\n"
            "       Make sure you are using a Page Access Token (not a User Token)\n"
            "       or that your app has pages_show_list permission granted."
        )

    if len(pages) == 1:
        return pages[0]["id"], pages[0]["name"]

    # Multiple pages — list them so the user can pick with --page-id.
    page_list = "\n".join(
        f"  {p['id']}  {p['name']}" for p in pages
    )
    sys.exit(
        f"ERROR: Token has access to {len(pages)} pages. "
        "Use --page-id to specify which one:\n" + page_list
    )


def fetch_posts(token: str, page_id: str):
    """
    Fetch all posts from the given Facebook Page via the Graph API and match
    them against the NJ Town caption pattern.

    The first page is fetched via _graph_get(); subsequent pages follow the
    raw 'paging.next' URL returned by the API (which already contains the
    cursor and all required parameters).

    Returns:
        total      — total posts fetched from the API
        matched    — list of dicts: {visit_number, is_bonus, post_url, caption_line}
        unmatched  — count of posts with no "NJ Town #N" pattern
    """
    matched   = []
    unmatched = 0
    total     = 0
    next_url  = None   # raw pagination URL; None on the first iteration

    while True:
        # ── Fetch one page ────────────────────────────────────────────────────
        if next_url:
            # Subsequent pages: paging.next already encodes all params + cursor
            resp = requests.get(next_url, timeout=30)
            try:
                data = resp.json()
            except ValueError:
                sys.exit(
                    f"ERROR: Non-JSON response during pagination (HTTP {resp.status_code})."
                )
            if not resp.ok or "error" in data:
                err = data.get("error", {})
                sys.exit(
                    f"ERROR: Graph API error during pagination "
                    f"(HTTP {resp.status_code}).\n"
                    f"       Message : {err.get('message', 'unknown error')}"
                )
        else:
            # First page via the shared helper (handles error reporting)
            data = _graph_get(
                f"{page_id}/posts",
                {
                    "access_token": token,
                    "fields":       "message,permalink_url",
                    "limit":        _PAGE_SIZE,
                },
            )

        # ── Process posts in this page ────────────────────────────────────────
        posts  = data.get("data", [])
        total += len(posts)

        for post in posts:
            message = (post.get("message") or "").strip()
            url     = (post.get("permalink_url") or "").strip()

            if not message or not url:
                continue

            line = _first_line(message)
            m    = _TOWN_RE.match(line)
            if not m:
                unmatched += 1
                continue

            visit_num = int(m.group(1))
            is_bonus  = bool(_BONUS_RE.search(line))

            matched.append({
                "visit_number": visit_num,
                "is_bonus":     is_bonus,
                "post_url":     url,
                "caption_line": line,
            })

        # ── Advance to next page or stop ──────────────────────────────────────
        next_url = data.get("paging", {}).get("next")
        if not next_url:
            break

    return total, matched, unmatched


def load_excel(excel_path: Path):
    """
    Open the workbook and return:
        wb            — the Workbook object
        ws            — the active worksheet (Data sheet)
        col           — dict: field_name -> 1-based column index
        visit_to_row  — dict: visit_number (int) -> row index (int)
        visit_to_name — dict: visit_number (int) -> town name (str)

    Row 1 = group labels (skipped); Row 2 = field names; Rows 3+ = data.
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Read field names from row 2
    col = {}
    for cell in ws[2]:
        if cell.value is not None:
            col[str(cell.value).strip()] = cell.column

    required = {
        "visitNumber", "name",
        "facebook1_label", "facebook1_url",
        "facebook2_label", "facebook2_url",
    }
    missing = sorted(required - col.keys())
    if missing:
        sys.exit(
            f"ERROR: Required columns not found in row 2 of the Excel file: {missing}\n"
            "       Make sure you are pointing at ejc-data-tracker.xlsx (not the template)."
        )

    visit_to_row  = {}
    visit_to_name = {}
    for row in ws.iter_rows(min_row=3):
        v_cell = row[col["visitNumber"] - 1]
        n_cell = row[col["name"] - 1]
        if v_cell.value is None:
            continue
        try:
            vnum = int(v_cell.value)
        except (TypeError, ValueError):
            continue
        visit_to_row[vnum]  = v_cell.row
        visit_to_name[vnum] = str(n_cell.value or "").strip()

    return wb, ws, col, visit_to_row, visit_to_name


def write_to_excel(ws, col, visit_to_row, visit_to_name, matched):
    """
    Write matched Facebook posts into the appropriate cells of the worksheet.

    Returns:
        written          — number of label+url pairs written
        skipped_full     — skipped because the target slot was already filled
        skipped_missing  — skipped because the visit number was not found in Excel
    """
    written         = 0
    skipped_full    = 0
    skipped_missing = 0

    for post in matched:
        vnum     = post["visit_number"]
        is_bonus = post["is_bonus"]
        url      = post["post_url"]

        if vnum not in visit_to_row:
            print(f"  WARNING: No row found for visit #{vnum} — skipping {url}")
            skipped_missing += 1
            continue

        row_idx   = visit_to_row[vnum]
        town_name = visit_to_name.get(vnum, f"Town #{vnum}")

        # Determine target slot and build normalized label
        if is_bonus:
            label     = f"NJ Town #{vnum}: Bonus Town Trivia"
            label_col = col["facebook2_label"]
            url_col   = col["facebook2_url"]
        else:
            label     = f"NJ Town #{vnum}: {town_name}"
            label_col = col["facebook1_label"]
            url_col   = col["facebook1_url"]

        existing_label = ws.cell(row=row_idx, column=label_col).value
        existing_url   = ws.cell(row=row_idx, column=url_col).value

        if existing_label or existing_url:
            slot = "facebook2" if is_bonus else "facebook1"
            print(f"  SKIP (already filled): visit #{vnum} [{slot}] — {url}")
            skipped_full += 1
            continue

        ws.cell(row=row_idx, column=label_col).value = label
        ws.cell(row=row_idx, column=url_col).value   = url
        written += 1

    return written, skipped_full, skipped_missing


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description=(
            "Fetch Facebook Page posts via the Graph API and import URLs "
            "into data/ejc-data-tracker.xlsx."
        )
    )
    parser.add_argument(
        "--token",
        metavar="ACCESS_TOKEN",
        default="",
        help=(
            "Facebook Page Access Token. "
            "Can also be set via the FB_PAGE_TOKEN environment variable "
            "(env var takes precedence)."
        ),
    )
    parser.add_argument(
        "--page-id",
        metavar="PAGE_ID",
        default="",
        help=(
            "Facebook Page ID (numeric). "
            "Auto-detected from the token via /me/accounts if not provided."
        ),
    )
    parser.add_argument(
        "--excel",
        type=Path,
        default=EXCEL_PATH,
        metavar="PATH",
        help=f"Path to the Excel tracker file (default: {EXCEL_PATH})",
    )
    args = parser.parse_args()

    excel_path = args.excel.resolve()

    # ── Resolve token (env var takes precedence over CLI arg) ─────────────────
    token = os.environ.get("FB_PAGE_TOKEN", "").strip() or args.token.strip()
    if not token:
        sys.exit(
            "ERROR: No access token provided.\n"
            "       Set the FB_PAGE_TOKEN environment variable or use --token YOUR_TOKEN.\n\n"
            "How to get a Page Access Token:\n"
            "  1. Go to https://developers.facebook.com/tools/explorer\n"
            "  2. Create or select a Meta App\n"
            "  3. Click 'Get User Token' → grant pages_read_engagement + pages_show_list\n"
            "  4. Call GET /me/accounts and copy the access_token for eatjerseychallenge"
        )

    # ── Validate Excel path ───────────────────────────────────────────────────
    if not excel_path.exists():
        sys.exit(f"ERROR: Excel file not found: {excel_path}")

    # ── Resolve page ID ───────────────────────────────────────────────────────
    page_id = args.page_id.strip()
    if not page_id:
        print("Detecting page : calling /me/accounts …")
        page_id, page_name = get_page_id(token)
        print(f"Page detected  : {page_name} (ID: {page_id})")
    else:
        print(f"Page ID        : {page_id}")

    # ── Fetch posts from Graph API ────────────────────────────────────────────
    print(f"Fetching posts : GET /{page_id}/posts …")
    total, matched, unmatched_count = fetch_posts(token, page_id)

    print(f"Posts fetched  : {total}")
    print(f"  Matched 'NJ Town #N' : {len(matched)}")
    if unmatched_count:
        print(f"  No town pattern (skipped) : {unmatched_count}")

    if not matched:
        print("\nNo matching posts found. Nothing to write.")
        return

    # ── Load Excel ────────────────────────────────────────────────────────────
    print(f"Reading        : {excel_path}")
    wb, ws, col, visit_to_row, visit_to_name = load_excel(excel_path)

    # ── Backup before any writes ──────────────────────────────────────────────
    _make_backup(excel_path)

    # ── Write updates ─────────────────────────────────────────────────────────
    written, skipped_full, skipped_missing = write_to_excel(
        ws, col, visit_to_row, visit_to_name, matched
    )

    if written == 0:
        print(
            "\nNo new data to write — "
            "all matched posts already have entries or had no matching visit number."
        )
        return

    wb.save(excel_path)
    print(f"\nSaved          : {excel_path}")
    print(f"Written        : {written} post(s)")
    if skipped_full:
        print(f"Already filled : {skipped_full} slot(s) skipped")
    if skipped_missing:
        print(f"Not in Excel   : {skipped_missing} visit number(s) not found")


if __name__ == "__main__":
    main()
