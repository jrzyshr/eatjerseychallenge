"""
scripts/generate-excel-template.py

Generates data/ejc-data-template.xlsx — the Excel spreadsheet used to maintain
all municipality visit data for the Eat Jersey Challenge.

The Summary and county sheets use live Excel formulas referencing the Data sheet,
so they update automatically as you edit the Data sheet — no need to re-run this
script during normal use.

Re-run this script only when:
  - You run npm run import to bulk-load data from a CSV
  - The column structure changes

After editing the Data sheet, export to CSV and run:
    npm run import -- path/to/exported.csv

Usage:
    /Users/plaudati/EJCWeb/eatjerseychallenge/.venv/bin/python scripts/generate-excel-template.py
"""

import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT = Path(__file__).parent.parent / "data" / "ejc-data-template.xlsx"

# ── Column definitions ────────────────────────────────────────────────────────
# Each tuple: (header, width, note)
COLUMNS = [
    # ── Identity ──
    ("name",             22, "Municipality name, e.g. Absecon  [REQUIRED if no geoid]"),
    ("county",           16, "County name, e.g. Atlantic  [used to disambiguate duplicate names]"),
    ("geoid",            16, "10-digit GEOID from GeoJSON  [optional — preferred over name lookup]"),
    ("townType",         14, "Auto-derived from GeoJSON; one of: city | borough | township | town | village"),

    # ── Visit data ──
    ("status",           18, "One of: unvisited | visited | queued | pre-challenge"),
    ("visitNumber",      14, "Order visited on the journey, 1–564"),
    ("restaurantName",   28, "Primary restaurant / business visited"),
    ("dateVisited",      14, "Date visited, formatted YYYY-MM-DD"),

    # ── Restaurant link ──
    ("restaurant_label", 26, "Display text for restaurant link  (defaults to restaurantName)"),
    ("restaurant_url",   40, "URL to restaurant website or social page"),

    # ── Wikipedia ──
    ("wikipedia_url",    40, "URL to the municipality's Wikipedia article"),

    # ── Instagram (up to 2 posts) ──
    ("instagram1_label", 26, "Description of first Instagram post, e.g. 'Restaurant Name'"),
    ("instagram1_url",   40, "URL to first Instagram post"),
    ("instagram2_label", 26, "Description of second Instagram post, e.g. 'Bonus Town Trivia'"),
    ("instagram2_url",   40, "URL to second Instagram post"),

    # ── TikTok (up to 2 posts) ──
    ("tiktok1_label",    26, "Description of first TikTok post"),
    ("tiktok1_url",      40, "URL to first TikTok post"),
    ("tiktok2_label",    26, "Description of second TikTok post"),
    ("tiktok2_url",      40, "URL to second TikTok post"),

    # ── YouTube (up to 2 posts) ──
    ("youtube1_label",   26, "Description of first YouTube video"),
    ("youtube1_url",     40, "URL to first YouTube video"),
    ("youtube2_label",   26, "Description of second YouTube video"),
    ("youtube2_url",     40, "URL to second YouTube video"),

    # ── Threads (up to 2 posts) ──
    ("threads1_label",   26, "Description of first Threads post"),
    ("threads1_url",     40, "URL to first Threads post"),
    ("threads2_label",   26, "Description of second Threads post"),
    ("threads2_url",     40, "URL to second Threads post"),

    # ── Bluesky (up to 2 posts) ──
    ("bluesky1_label",   26, "Description of first Bluesky post"),
    ("bluesky1_url",     40, "URL to first Bluesky post"),
    ("bluesky2_label",   26, "Description of second Bluesky post"),
    ("bluesky2_url",     40, "URL to second Bluesky post"),

    # ── Facebook (up to 2 posts) ──
    ("facebook1_label",  26, "Description of first Facebook post, e.g. 'Restaurant Name'"),
    ("facebook1_url",    40, "URL to first Facebook post"),
    ("facebook2_label",  26, "Description of second Facebook post"),
    ("facebook2_url",    40, "URL to second Facebook post"),

    # ── Extra overflow slots (3) ──
    ("extra1_category",  18, "Category: restaurant | wikipedia | social | more | custom text"),
    ("extra1_platform",  16, "Platform name if category=social, e.g. Instagram"),
    ("extra1_label",     26, "Display text for this link"),
    ("extra1_url",       40, "URL"),

    ("extra2_category",  18, "Category for second extra link"),
    ("extra2_platform",  16, "Platform name if category=social"),
    ("extra2_label",     26, "Display text for second extra link"),
    ("extra2_url",       40, "URL"),

    ("extra3_category",  18, "Category for third extra link"),
    ("extra3_platform",  16, "Platform name if category=social"),
    ("extra3_label",     26, "Display text for third extra link"),
    ("extra3_url",       40, "URL"),
]

# ── Column group colour bands (applied to header row) ─────────────────────────
GROUP_COLORS = {
    "identity":    "D9E1F2",   # blue-grey
    "visit":       "E2EFDA",   # green
    "restaurant":  "FFF2CC",   # yellow
    "wikipedia":   "DDEBF7",   # light blue
    "instagram":   "FCE4EC",   # pink
    "tiktok":      "F3E5F5",   # purple
    "youtube":     "FFEBEE",   # red
    "threads":     "E8F5E9",   # light green
    "bluesky":     "E3F2FD",   # sky blue
    "facebook":    "DDEEFF",   # blue (Facebook brand)
    "extra":       "F5F5F5",   # grey
}

def column_group(col_name):
    if col_name in ("name", "county", "geoid", "townType"):
        return "identity"
    if col_name in ("status", "visitNumber", "restaurantName", "dateVisited"):
        return "visit"
    if col_name.startswith("restaurant"):
        return "restaurant"
    if col_name.startswith("wikipedia"):
        return "wikipedia"
    if col_name.startswith("instagram"):
        return "instagram"
    if col_name.startswith("tiktok"):
        return "tiktok"
    if col_name.startswith("youtube"):
        return "youtube"
    if col_name.startswith("threads"):
        return "threads"
    if col_name.startswith("bluesky"):
        return "bluesky"
    if col_name.startswith("facebook"):
        return "facebook"
    if col_name.startswith("extra"):
        return "extra"
    return "extra"

# ── Style helpers ─────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)

def build():
    # ── Load municipality data ────────────────────────────────────────────────
    json_path = Path(__file__).parent.parent / 'data' / 'municipalities.json'
    with open(json_path, 'r', encoding='utf-8') as f:
        muni_data = json.load(f)

    # Sort all towns alphabetically; keep geoid alongside each record
    all_towns   = sorted(muni_data.items(), key=lambda kv: kv[1].get('name', ''))
    # all_towns   = [(geoid, data), ...]

    # Group by county (preserving alphabetical sort within each county)
    county_towns = {}
    for geoid, town in all_towns:
        county = town.get('county', 'Unknown')
        county_towns.setdefault(county, []).append((geoid, town))

    def is_visited(town):
        return town.get('status', 'unvisited') != 'unvisited'

    total_towns   = len(all_towns)
    total_visited = sum(1 for _, t in all_towns if is_visited(t))

    headers = [c[0] for c in COLUMNS]

    STATUS_DV_FORMULA = '"unvisited,visited,queued,pre-challenge"'
    status_col        = headers.index('status') + 1

    GROUP_LABELS = {
        "identity":   "MUNICIPALITY",
        "visit":      "VISIT DATA",
        "restaurant": "RESTAURANT LINK",
        "wikipedia":  "WIKIPEDIA",
        "instagram":  "INSTAGRAM",
        "tiktok":     "TIKTOK",
        "youtube":    "YOUTUBE",
        "threads":    "THREADS",
        "bluesky":    "BLUESKY",
        "facebook":   "FACEBOOK",
        "extra":      "EXTRA / OVERFLOW LINKS",
    }

    wb = Workbook()

    # ── Shared helpers ────────────────────────────────────────────────────────
    def write_col_headers(ws, start_row):
        """Write group-label row and column-header row starting at start_row."""
        grow = start_row
        hrow = start_row + 1

        # Build group spans
        prev_group = None
        span_start = 1
        spans      = []
        for i, col in enumerate(headers, start=1):
            g = column_group(col)
            if g != prev_group:
                if prev_group is not None:
                    spans.append((prev_group, span_start, i - 1))
                prev_group = g
                span_start = i
        spans.append((prev_group, span_start, len(headers)))

        for (group, start, end) in spans:
            cell = ws.cell(row=grow, column=start,
                           value=GROUP_LABELS.get(group, group.upper()))
            cell.font      = Font(bold=True, size=9, color="444444")
            cell.fill      = fill(GROUP_COLORS[group])
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = thin_border()
            if end > start:
                ws.merge_cells(start_row=grow, start_column=start,
                               end_row=grow, end_column=end)
        ws.row_dimensions[grow].height = 18

        for col_idx, (col_name, width, _) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=hrow, column=col_idx, value=col_name)
            cell.font      = Font(bold=True, size=10, color="1a1a1a")
            cell.fill      = fill(GROUP_COLORS[column_group(col_name)])
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            cell.border    = thin_border()
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[hrow].height = 30

    def add_status_dv(ws, first_data_row, last_row=600):
        dv = DataValidation(
            type="list", formula1=STATUS_DV_FORMULA,
            allow_blank=True, showDropDown=False, showErrorMessage=True,
            errorTitle="Invalid status",
            error="Must be one of: unvisited, visited, queued, pre-challenge"
        )
        col_letter = get_column_letter(status_col)
        dv.sqref   = f"{col_letter}{first_data_row}:{col_letter}{last_row}"
        ws.add_data_validation(dv)

    def write_data_row(ws, row_idx, geoid, town, shade):
        row_data = {
            'name':           town.get('name', ''),
            'county':         town.get('county', ''),
            'geoid':          geoid,
            'townType':       town.get('townType') or '',
            'status':         town.get('status', 'unvisited'),
            'visitNumber':    '' if town.get('visitNumber') is None else town['visitNumber'],
            'restaurantName': town.get('restaurantName') or '',
            'dateVisited':    town.get('dateVisited') or '',
        }
        for col_idx, col_name in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx,
                           value=row_data.get(col_name, ''))
            cell.fill      = fill(shade)
            cell.border    = thin_border()
            cell.font      = Font(size=10)
            cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[row_idx].height = 16

    # ── DATA sheet ────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title        = "Data"
    ws.freeze_panes = "D3"   # freeze cols A-C + rows 1-2 (group labels + headers)

    write_col_headers(ws, start_row=1)
    add_status_dv(ws, first_data_row=3)

    for offset, (geoid, town) in enumerate(all_towns):
        shade = "FFFFFF" if offset % 2 == 0 else "F7F7F7"
        write_data_row(ws, 3 + offset, geoid, town, shade)

    # ── SUMMARY sheet (formula-driven — auto-updates as Data sheet changes) ───
    # Data sheet: status is col E (index 5), county is col B (index 2)
    # Data rows start at row 3, we count up to row 602 to cover all 564 towns
    STATUS_COL   = get_column_letter(headers.index('status') + 1)   # E
    COUNTY_COL   = get_column_letter(headers.index('county') + 1)   # B
    DATA_FIRST   = 3
    DATA_LAST    = 602   # generous upper bound
    status_range = f"Data!{STATUS_COL}{DATA_FIRST}:{STATUS_COL}{DATA_LAST}"
    county_range = f"Data!{COUNTY_COL}{DATA_FIRST}:{COUNTY_COL}{DATA_LAST}"

    ws_s = wb.create_sheet("Summary")
    ws_s.column_dimensions["A"].width = 18
    for col in ["B", "C", "D", "E"]:
        ws_s.column_dimensions[col].width = 14

    # Title bar
    title = ws_s.cell(row=1, column=1,
                      value="EAT JERSEY CHALLENGE — Statewide Progress")
    title.font      = Font(bold=True, size=14, color="FFFFFF")
    title.fill      = fill("1e2d47")
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws_s.merge_cells("A1:E1")
    ws_s.row_dimensions[1].height = 28

    # Overall stat boxes — formula-driven
    # Total = COUNTA(status range excluding blanks) via COUNTA on name col
    name_range    = f"Data!{get_column_letter(headers.index('name')+1)}{DATA_FIRST}:{get_column_letter(headers.index('name')+1)}{DATA_LAST}"
    f_total   = f"=COUNTA({name_range})"
    f_visited = f'=COUNTIF({status_range},"visited")+COUNTIF({status_range},"queued")+COUNTIF({status_range},"pre-challenge")'
    f_togo    = f"=B3-C3"
    f_pct     = f'=IF(B3>0,TEXT(C3/B3,"0.0%"),"0.0%")'

    stat_boxes = [
        ("Total Towns", f_total,   "D9E1F2"),
        ("Visited",     f_visited, "E2EFDA"),
        ("To Go",       f_togo,    "FFEBEE"),
        ("% Complete",  f_pct,     "FFF2CC"),
    ]
    for col_idx, (label, formula, colour) in enumerate(stat_boxes, start=1):
        lc = ws_s.cell(row=2, column=col_idx, value=label)
        vc = ws_s.cell(row=3, column=col_idx, value=formula)
        for cell, bold, size in [(lc, True, 10), (vc, True, 14)]:
            cell.font      = Font(bold=bold, size=size)
            cell.fill      = fill(colour)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = thin_border()
    ws_s.row_dimensions[2].height = 18
    ws_s.row_dimensions[3].height = 24
    ws_s.row_dimensions[4].height = 10   # spacer

    # County table header (row 5)
    for col_idx, heading in enumerate(
            ["County", "Total Towns", "Visited", "To Go", "% Complete"],
            start=1):
        cell = ws_s.cell(row=5, column=col_idx, value=heading)
        cell.font      = Font(bold=True, size=10, color="FFFFFF")
        cell.fill      = fill("1e2d47")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin_border()
    ws_s.row_dimensions[5].height = 18

    # County data rows — COUNTIFS formulas reference Data sheet
    for offset, county in enumerate(sorted(county_towns.keys())):
        shade   = "FFFFFF" if offset % 2 == 0 else "F5F5F5"
        row_idx = 6 + offset
        f_ctotal   = f'=COUNTIF({county_range},"{county}")'
        f_cvisited = (f'=COUNTIFS({county_range},"{county}",{status_range},"visited")'
                      f'+COUNTIFS({county_range},"{county}",{status_range},"queued")'
                      f'+COUNTIFS({county_range},"{county}",{status_range},"pre-challenge")')
        f_ctogo    = f"=B{row_idx}-C{row_idx}"
        f_cpct     = f'=IF(B{row_idx}>0,TEXT(C{row_idx}/B{row_idx},"0.0%"),"0.0%")'
        for col_idx, val in enumerate(
                [county, f_ctotal, f_cvisited, f_ctogo, f_cpct],
                start=1):
            cell = ws_s.cell(row=row_idx, column=col_idx, value=val)
            cell.fill      = fill(shade)
            cell.border    = thin_border()
            cell.font      = Font(size=10)
            cell.alignment = Alignment(
                horizontal="left" if col_idx == 1 else "center",
                vertical="center"
            )
        ws_s.row_dimensions[row_idx].height = 15

    # ── COUNTY sheets (formula-driven via FILTER — auto-updates) ─────────────
    # FILTER spills all Data rows where county column matches this county.
    # Covers all 47 columns via the full Data range A3:AV602.
    last_data_col   = get_column_letter(len(headers))  # e.g. AV
    full_data_range = f"Data!A{DATA_FIRST}:{last_data_col}{DATA_LAST}"
    county_col_full = f"Data!{COUNTY_COL}{DATA_FIRST}:{COUNTY_COL}{DATA_LAST}"

    for county in sorted(county_towns.keys()):
        towns     = county_towns[county]
        c_total   = len(towns)

        ws_c = wb.create_sheet(county[:31])

        # Apply column widths to match Data sheet
        for col_idx, (_, width, _) in enumerate(COLUMNS, start=1):
            ws_c.column_dimensions[get_column_letter(col_idx)].width = width

        # Row 1 — county title bar
        title_cell = ws_c.cell(row=1, column=1, value=f"{county} County")
        title_cell.font      = Font(bold=True, size=13, color="FFFFFF")
        title_cell.fill      = fill("1e2d47")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_c.merge_cells(start_row=1, start_column=1,
                         end_row=1, end_column=len(headers))
        ws_c.row_dimensions[1].height = 24

        # Row 2 — summary stat boxes (COUNTIFS formulas)
        f_ctotal   = f'=COUNTIF({county_col_full},"{county}")'
        f_cvisited = (f'=COUNTIFS({county_col_full},"{county}",{status_range},"visited")'
                      f'+COUNTIFS({county_col_full},"{county}",{status_range},"queued")'
                      f'+COUNTIFS({county_col_full},"{county}",{status_range},"pre-challenge")')
        f_ctogo    = "=B2-D2"
        county_stats = [
            ("Total Towns", f_ctotal,   "D9E1F2"),
            ("Visited",     f_cvisited, "E2EFDA"),
            ("To Go",       f_ctogo,    "FFEBEE"),
        ]
        col_cursor = 1
        for label, formula, colour in county_stats:
            lc = ws_c.cell(row=2, column=col_cursor,     value=label)
            vc = ws_c.cell(row=2, column=col_cursor + 1, value=formula)
            for cell, bold in [(lc, True), (vc, False)]:
                cell.font      = Font(bold=bold, size=10)
                cell.fill      = fill(colour)
                cell.border    = thin_border()
                cell.alignment = Alignment(horizontal="center", vertical="center")
            col_cursor += 2
        ws_c.row_dimensions[2].height = 20

        # Row 3 — spacer
        ws_c.row_dimensions[3].height = 8

        # Rows 4-5 — column headers (group row + field name row)
        write_col_headers(ws_c, start_row=4)

        # Row 6 — FILTER formula (spills down automatically in Excel 365)
        filter_formula = (f'=IFERROR(FILTER({full_data_range},'
                          f'{county_col_full}="{county}"),"")')
        ws_c.cell(row=6, column=1, value=filter_formula)
        ws_c.row_dimensions[6].height = 16

        # Freeze: col A-C + rows 1-5
        ws_c.freeze_panes = "D6"

    # ── LEGEND sheet ─────────────────────────────────────────────────────────
    wl = wb.create_sheet("Legend")
    wl.column_dimensions["A"].width = 26
    wl.column_dimensions["B"].width = 16
    wl.column_dimensions["C"].width = 65

    def legend_header(row, text):
        cell = wl.cell(row=row, column=1, value=text)
        cell.font      = Font(bold=True, size=11, color="FFFFFF")
        cell.fill      = fill("1e2d47")
        cell.alignment = Alignment(vertical="center")
        wl.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        wl.row_dimensions[row].height = 20

    def legend_row(row, col, width, note):
        for c, v in [(1, col), (2, f"{width} chars"), (3, note)]:
            cell = wl.cell(row=row, column=c, value=v)
            cell.fill      = fill(GROUP_COLORS[column_group(col)])
            cell.border    = thin_border()
            cell.font      = Font(size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=(c == 3))
        wl.row_dimensions[row].height = 14

    wl.row_dimensions[1].height = 14
    lrow = 2
    legend_header(lrow, "Column Reference — Eat Jersey Challenge Data Sheet"); lrow += 1

    lrow += 1
    legend_header(lrow, "Status Values"); lrow += 1
    for val, desc in [
        ("unvisited",     "Town not yet visited"),
        ("visited",       "Visited AND social posts are published"),
        ("queued",        "Visited, but posts not yet published (shows as visited on map)"),
        ("pre-challenge", "Visited before the challenge started (shows as visited on map)"),
    ]:
        wl.cell(row=lrow, column=1, value=val).font  = Font(bold=True, size=10)
        wl.cell(row=lrow, column=3, value=desc).font = Font(size=10)
        for c in range(1, 4):
            wl.cell(row=lrow, column=c).border = thin_border()
            wl.cell(row=lrow, column=c).fill   = fill("F5F5F5")
        lrow += 1

    lrow += 1
    legend_header(lrow, "Sheet Guide"); lrow += 1
    sheet_notes = [
        ("Data",          "Master data sheet — this is the one you export to CSV for import"),
        ("Summary",       "Statewide progress: totals + per-county breakdown. AUTO-UPDATES as you edit the Data sheet."),
        ("<County name>", "One sheet per NJ county. Row data auto-updates via FILTER. Stats auto-update via COUNTIFS."),
        ("Legend",        "This reference sheet"),
        ("Note",          "Re-run gen-template only after npm run import (bulk CSV load) or if column structure changes"),
        ("Note",          "To export: have the Data sheet active → File → Save As → CSV"),
    ]
    for step, text in sheet_notes:
        wl.cell(row=lrow, column=1, value=step).font  = Font(bold=True, size=10)
        wl.cell(row=lrow, column=3, value=text).font  = Font(size=10)
        for c in range(1, 4):
            wl.cell(row=lrow, column=c).border = thin_border()
            wl.cell(row=lrow, column=c).fill   = fill("E3F2FD")
        wl.row_dimensions[lrow].height = 16
        lrow += 1

    lrow += 1
    legend_header(lrow, "All Columns"); lrow += 1
    for (col, width, note) in COLUMNS:
        legend_row(lrow, col, width, note)
        lrow += 1

    lrow += 1
    legend_header(lrow, "Export to JSON"); lrow += 1
    instructions = [
        ("Step 1", "Edit the Data sheet directly in Excel — Summary and county sheets update automatically"),
        ("Step 2", "To sync to the website: make the Data sheet active → File → Save As → CSV"),
        ("Step 3", "In your terminal: npm run import -- path/to/your-file.csv"),
        ("Step 4", "Commit and push municipalities.json to deploy to the website"),
        ("Notes",  "Columns left blank in CSV are ignored — existing values in municipalities.json are not erased"),
        ("Notes",  "The extra sheets (county + summary) do NOT affect CSV export — only the active sheet exports"),
        ("Notes",  "County sheets use Excel 365 FILTER function — requires Microsoft 365 or Excel 2021+"),
        ("Notes",  "Re-run gen-template only if municipalities.json was edited outside of Excel, or columns change"),
    ]
    for (step, text) in instructions:
        wl.cell(row=lrow, column=1, value=step).font  = Font(bold=True, size=10)
        wl.cell(row=lrow, column=3, value=text).font  = Font(size=10)
        for c in range(1, 4):
            wl.cell(row=lrow, column=c).border = thin_border()
            wl.cell(row=lrow, column=c).fill   = fill("FFFDE7")
        wl.row_dimensions[lrow].height = 16
        lrow += 1

    wb.save(OUTPUT)
    print(f"Generated: {OUTPUT}")
    print(f"  Sheets   : Data, Summary, {len(county_towns)} county sheets, Legend")
    print(f"  Towns    : {total_towns} total, {total_visited} visited")
    print( "  Open the Legend sheet for column reference and export instructions.")

if __name__ == "__main__":
    build()
