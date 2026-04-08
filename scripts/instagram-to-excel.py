"""scripts/instagram-to-excel.py

Reads a PostFox "Export IG Posts" JSON file and writes Instagram post URLs
and labels into the appropriate columns of data/ejc-data-tracker.xlsx.

PostFox (Chrome extension) exports a flat JSON array of post objects.
The relevant fields used by this script are:
    "Caption"   — full caption text (newline-separated paragraphs)
    "Post URL"  — permalink, e.g. https://www.instagram.com/p/DQo0IldESTD/
    "User Name" — Instagram username of the post owner

Posts are matched to towns by extracting the visit number from the first line
of the caption and looking it up against the visitNumber column in the Excel
file.  Only posts that follow the "NJ Town #N:" or "New Jersey Town N:" pattern
are processed; all others are silently skipped.

Slot assignment:
    "Bonus Town Trivia" in the caption first line → instagram2_label / instagram2_url
    All other matching posts                       → instagram1_label / instagram1_url

Labels are always normalized to:
    Main post:   "NJ Town #N: <Town Name>"        (town name from Excel, not caption)
    Bonus post:  "NJ Town #N: Bonus Town Trivia"

Cells that already contain data are never overwritten.
The extra1/extra2/extra3 columns are never touched.
A timestamped backup of the Excel file is created before any changes are saved.

Usage:
    .venv/bin/python scripts/instagram-to-excel.py data/IGPOSTS_USERS_eatjerseychallenge_62.json
    .venv/bin/python scripts/instagram-to-excel.py path/to/file.json --username eatjerseychallenge
    .venv/bin/python scripts/instagram-to-excel.py path/to/file.json --excel path/to/other.xlsx
"""

import argparse
import json
import re
import shutil
import sys
from datetime import date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit(
        "ERROR: openpyxl is not installed.\n"
        "       Activate the virtual environment and run: pip install openpyxl"
    )

# ── Repo paths ────────────────────────────────────────────────────────────────
_REPO_ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = _REPO_ROOT / "data" / "ejc-data-tracker.xlsx"

# ── Caption matching ──────────────────────────────────────────────────────────
# Matches the start of a caption first line like:
#   "NJ Town #93: Carteret"
#   "NJ Town 93: Carteret"
#   "New Jersey Town #93: Carteret"
#   "NJ Town #93: Carteret - Bonus Town Trivia"
# Captures the numeric visit number.
_TOWN_RE = re.compile(
    r"^(?:NJ|New\s+Jersey)\s+Town\s+#?(\d+)\s*[:\-,]",
    re.IGNORECASE,
)

# Detects "Bonus Town Trivia" anywhere in the first caption line.
_BONUS_RE = re.compile(r"Bonus\s+Town\s+Trivia", re.IGNORECASE)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _first_line(text: str) -> str:
    """Return the first non-empty line of a multi-line string."""
    for line in text.splitlines():
        stripped = line.strip()
        if stripped:
            return stripped
    return ""


def _detect_username(path: Path) -> str:
    """
    Extract the Instagram username from a PostFox filename like:
        IGPOSTS_USERS_{username}_{n}.json
    Returns an empty string if the filename does not match the pattern.
    """
    m = re.match(r"(?i)IGPOSTS_USERS_(.+?)_\d+\.json$", path.name)
    return m.group(1) if m else ""


def _make_backup(excel_path: Path) -> Path:
    """
    Copy the Excel file to a date-stamped backup path.
    Does not overwrite a backup already created earlier the same day.
    Returns the backup path.
    """
    stamp = date.today().strftime("%Y%m%d")
    backup = excel_path.with_name(f"ejc-data-tracker-backup-{stamp}.xlsx")
    if not backup.exists():
        shutil.copy2(excel_path, backup)
        print(f"Backup created : {backup}")
    else:
        print(f"Backup exists  : {backup}  (not overwritten)")
    return backup


# ── Core logic ────────────────────────────────────────────────────────────────

def parse_posts(json_path: Path, username: str):
    """
    Load the PostFox JSON export and return:
        total        — total number of entries in the file
        user_count   — entries belonging to the target username
        matched      — list of dicts: {visit_number, is_bonus, post_url, caption_line}
        unmatched    — count of user posts with no "NJ Town #N" pattern
    """
    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)

    if not isinstance(data, list):
        sys.exit(
            f"ERROR: Expected a JSON array, got {type(data).__name__}.\n"
            "       Make sure you are using a PostFox JSON export file."
        )

    total      = len(data)
    user_count = 0
    matched    = []
    unmatched  = 0

    for entry in data:
        if not isinstance(entry, dict):
            continue
        entry_user = (entry.get("User Name") or "").strip()
        if entry_user.lower() != username.lower():
            continue

        user_count += 1

        caption  = (entry.get("Caption") or "").strip()
        post_url = (entry.get("Post URL") or "").strip()
        if not caption or not post_url:
            continue

        line = _first_line(caption)
        m    = _TOWN_RE.match(line)
        if not m:
            unmatched += 1
            continue

        visit_num = int(m.group(1))
        is_bonus  = bool(_BONUS_RE.search(line))

        matched.append({
            "visit_number": visit_num,
            "is_bonus":     is_bonus,
            "post_url":     post_url,
            "caption_line": line,
        })

    return total, user_count, matched, unmatched


def load_excel(excel_path: Path):
    """
    Open the workbook and return:
        wb            — the Workbook object
        ws            — the active worksheet (Data sheet)
        col           — dict: field_name -> 1-based column index
        visit_to_row  — dict: visit_number (int) -> row index (int)
        visit_to_name — dict: visit_number (int) -> town name (str)

    Row 1 = group labels (skipped); Row 2 = field names; Rows 3+ = data.
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Read field names from row 2
    col = {}
    for cell in ws[2]:
        if cell.value is not None:
            col[str(cell.value).strip()] = cell.column

    required = {
        "visitNumber", "name",
        "instagram1_label", "instagram1_url",
        "instagram2_label", "instagram2_url",
    }
    missing = sorted(required - col.keys())
    if missing:
        sys.exit(
            f"ERROR: Required columns not found in row 2 of the Excel file: {missing}\n"
            "       Make sure you are pointing at ejc-data-tracker.xlsx (not the template)."
        )

    visit_to_row  = {}
    visit_to_name = {}
    for row in ws.iter_rows(min_row=3):
        v_cell = row[col["visitNumber"] - 1]
        n_cell = row[col["name"] - 1]
        if v_cell.value is None:
            continue
        try:
            vnum = int(v_cell.value)
        except (TypeError, ValueError):
            continue
        visit_to_row[vnum]  = v_cell.row
        visit_to_name[vnum] = str(n_cell.value or "").strip()

    return wb, ws, col, visit_to_row, visit_to_name


def write_to_excel(ws, col, visit_to_row, visit_to_name, matched):
    """
    Write matched Instagram posts into the appropriate cells of the worksheet.

    Returns:
        written          — number of label+url pairs written
        skipped_full     — skipped because the target slot was already filled
        skipped_missing  — skipped because the visit number was not found in Excel
    """
    written         = 0
    skipped_full    = 0
    skipped_missing = 0

    for post in matched:
        vnum     = post["visit_number"]
        is_bonus = post["is_bonus"]
        url      = post["post_url"]

        if vnum not in visit_to_row:
            print(f"  WARNING: No row found for visit #{vnum} — skipping {url}")
            skipped_missing += 1
            continue

        row_idx   = visit_to_row[vnum]
        town_name = visit_to_name.get(vnum, f"Town #{vnum}")

        # Determine target slot and build normalized label
        if is_bonus:
            label     = f"NJ Town #{vnum}: Bonus Town Trivia"
            label_col = col["instagram2_label"]
            url_col   = col["instagram2_url"]
        else:
            label     = f"NJ Town #{vnum}: {town_name}"
            label_col = col["instagram1_label"]
            url_col   = col["instagram1_url"]

        existing_label = ws.cell(row=row_idx, column=label_col).value
        existing_url   = ws.cell(row=row_idx, column=url_col).value

        if existing_label or existing_url:
            slot = "instagram2" if is_bonus else "instagram1"
            print(f"  SKIP (already filled): visit #{vnum} [{slot}] — {url}")
            skipped_full += 1
            continue

        ws.cell(row=row_idx, column=label_col).value = label
        ws.cell(row=row_idx, column=url_col).value   = url
        written += 1

    return written, skipped_full, skipped_missing


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description=(
            "Import Instagram post URLs from a PostFox JSON export "
            "into data/ejc-data-tracker.xlsx."
        )
    )
    parser.add_argument(
        "json_file",
        type=Path,
        help=(
            "Path to the PostFox JSON export file "
            "(e.g. data/IGPOSTS_USERS_eatjerseychallenge_62.json)"
        ),
    )
    parser.add_argument(
        "--username",
        metavar="HANDLE",
        default="",
        help=(
            "Instagram username to filter posts by. "
            "Auto-detected from the filename if not provided."
        ),
    )
    parser.add_argument(
        "--excel",
        type=Path,
        default=EXCEL_PATH,
        metavar="PATH",
        help=f"Path to the Excel tracker file (default: {EXCEL_PATH})",
    )
    args = parser.parse_args()

    json_path  = args.json_file.resolve()
    excel_path = args.excel.resolve()

    # ── Validate inputs ───────────────────────────────────────────────────────
    if not json_path.exists():
        sys.exit(f"ERROR: File not found: {json_path}")
    if not excel_path.exists():
        sys.exit(f"ERROR: Excel file not found: {excel_path}")

    # ── Resolve username ──────────────────────────────────────────────────────
    username = args.username.strip()
    if not username:
        username = _detect_username(json_path)

    if not username:
        sys.exit(
            "ERROR: Could not auto-detect the Instagram username from the filename.\n"
            "       Use --username YOUR_HANDLE to specify it explicitly.\n"
            "       Expected filename pattern: IGPOSTS_USERS_{username}_{n}.json"
        )

    print(f"Username       : @{username}")

    # ── Parse PostFox JSON ────────────────────────────────────────────────────
    print(f"Reading        : {json_path}")
    total, user_count, matched, unmatched_count = parse_posts(json_path, username)

    print(f"Posts in file  : {total}")
    print(f"  @{username} posts : {user_count}")
    print(f"  Matched 'NJ Town #N' : {len(matched)}")
    if unmatched_count:
        print(f"  No town pattern (skipped) : {unmatched_count}")

    if not matched:
        print("\nNo matching posts found. Nothing to write.")
        return

    # ── Load Excel ────────────────────────────────────────────────────────────
    print(f"Reading        : {excel_path}")
    wb, ws, col, visit_to_row, visit_to_name = load_excel(excel_path)

    # ── Write updates ─────────────────────────────────────────────────────────
    written, skipped_full, skipped_missing = write_to_excel(
        ws, col, visit_to_row, visit_to_name, matched
    )

    if written == 0:
        print("\nNo new data to write — all matched posts already have entries or had no matching visit number.")
        return

    # ── Backup + save ─────────────────────────────────────────────────────────
    _make_backup(excel_path)
    wb.save(excel_path)
    print(f"Saved          : {excel_path}")

    # ── Summary ───────────────────────────────────────────────────────────────
    print()
    print("─" * 44)
    print(f"  Written                      : {written}")
    print(f"  Skipped (already filled)     : {skipped_full}")
    if skipped_missing:
        print(f"  Skipped (visit # not in Excel) : {skipped_missing}")
    print("─" * 44)
    print()
    print("Review the Excel file, then export to CSV and run:")
    print("  node scripts/excel-to-json.js path/to/export.csv")


if __name__ == "__main__":
    main()
